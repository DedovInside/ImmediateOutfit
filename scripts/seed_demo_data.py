"""
Загружает демонстрационный AARRR dataset из Excel и создаёт события для dashboard.

Запуск:
    python scripts/seed_demo_data.py

По умолчанию берёт файл из Telegram Desktop. Можно передать свой путь:
    python scripts/seed_demo_data.py "C:\\path\\ImmediateOutfit_AARRR_dataset.xlsx"
"""
from __future__ import annotations

import json
import sqlite3
import sys
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - запасной путь для окружения без openpyxl
    load_workbook = None

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from config import settings  # noqa: E402
from services.catalog import get_outfits  # noqa: E402
from services.storage import close_connection, init_db  # noqa: E402

DEFAULT_XLSX = Path(r"C:\Users\johnn\Downloads\Telegram Desktop\ImmediateOutfit_AARRR_dataset.xlsx")
MARKETING_SPEND_RUB = 2000
DEMO_USER_START = 900001
DEMO_USER_END = 900999

STEPS = ["gender", "occasion", "weather", "activity", "priority", "mood", "budget", "style"]
NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"Unsupported datetime format: {value!r}")


def _cell_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + ord(ch.upper()) - ord("A") + 1
    return index - 1


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ElementTree.fromstring(zf.read("xl/sharedStrings.xml"))
    values = []
    for item in root.findall("x:si", NS):
        parts = [node.text or "" for node in item.findall(".//x:t", NS)]
        values.append("".join(parts))
    return values


def _read_xlsx_sheet(path: Path, sheet_xml: str) -> list[list[object]]:
    with zipfile.ZipFile(path) as zf:
        shared = _read_shared_strings(zf)
        root = ElementTree.fromstring(zf.read(sheet_xml))
        table: list[list[object]] = []
        for row in root.findall(".//x:sheetData/x:row", NS):
            values: list[object] = []
            for cell in row.findall("x:c", NS):
                index = _cell_index(cell.attrib["r"])
                while len(values) <= index:
                    values.append(None)
                raw = cell.find("x:v", NS)
                value = raw.text if raw is not None else ""
                if cell.attrib.get("t") == "s":
                    values[index] = shared[int(value)]
                elif value == "":
                    values[index] = ""
                else:
                    number = float(value)
                    values[index] = int(number) if number.is_integer() else number
            table.append(values)
    return table


def _event(conn: sqlite3.Connection, user_id: int, name: str, when: datetime, payload: dict | None = None) -> None:
    payload_data = {"source": "demo_seed", **(payload or {})}
    conn.execute(
        """
        INSERT INTO events (user_id, event_name, payload_json, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (user_id, name, json.dumps(payload_data, ensure_ascii=False), when.isoformat()),
    )


def _seed_user_events(
    conn: sqlite3.Connection,
    row: dict,
    user_id: int,
    outfit_ids: list[str],
    outfit_snapshots: dict[str, str],
) -> None:
    start = row["first_contact_at"]
    end = start + timedelta(minutes=float(row["session_length_min"]))
    completed = int(row["scenario_completed"]) == 1
    scenarios = int(row["avg_scenarios"])
    bot_requests = int(row["bot_requests_count"])
    ai_count = int(row["ai_usage_count"])

    conn.execute(
        """
        INSERT OR REPLACE INTO users (user_id, username, first_name, last_name, created_at, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, row["user_code"].lower(), row["user_code"], "Demo", start.isoformat(), end.isoformat()),
    )

    gender = "female" if user_id % 5 else "male"
    style = ["casual", "base_minimal", "classic", "sport"][user_id % 4]
    budget = ["low", "medium", "medium", "medium"][user_id % 4]
    colors = ["черный", "белый"] if user_id % 2 else ["бежевый", "серый"]
    conn.execute(
        """
        INSERT OR REPLACE INTO profiles (
            user_id, gender, style, budget,
            preferred_colors, preferred_styles, disliked_items, key_items, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            gender,
            style,
            budget,
            json.dumps(colors, ensure_ascii=False),
            json.dumps([style], ensure_ascii=False),
            json.dumps(["каблуки"] if gender == "female" and user_id % 3 == 0 else [], ensure_ascii=False),
            json.dumps(["белая рубашка", "черные джинсы"], ensure_ascii=False),
            (start + timedelta(seconds=20)).isoformat(),
        ),
    )

    _event(conn, user_id, "bot_started", start)
    _event(conn, user_id, "profile_updated", start + timedelta(seconds=20), {"source": "demo_seed"})

    for scenario_index in range(max(1, scenarios)):
        scenario_start = start + timedelta(minutes=scenario_index)
        _event(conn, user_id, "quiz_started", scenario_start, {"mode": "standard" if scenario_index == 0 else "quick"})
        steps_to_write = STEPS if completed or scenario_index < scenarios - 1 else STEPS[:4]
        for step_index, step in enumerate(steps_to_write):
            _event(
                conn,
                user_id,
                "question_answered",
                scenario_start + timedelta(seconds=10 + step_index * 8),
                {"step": step, "value": "demo"},
            )
            if step == "weather":
                _event(conn, user_id, "weather_selected", scenario_start + timedelta(seconds=28), {"value": "mild"})

        if completed:
            _event(conn, user_id, "quiz_completed", scenario_start + timedelta(minutes=float(row["time_to_result_min"])))
            _event(
                conn,
                user_id,
                "results_viewed",
                scenario_start + timedelta(minutes=float(row["time_to_result_min"]), seconds=5),
                {"count": 2, "time_to_result_min": float(row["time_to_result_min"])},
            )
            if user_id % 2 == 0:
                outfit_id = outfit_ids[user_id % len(outfit_ids)]
                _event(conn, user_id, "outfit_saved", scenario_start + timedelta(minutes=3), {"outfit_id": outfit_id})
                conn.execute(
                    """
                    INSERT OR IGNORE INTO saved_outfits (user_id, outfit_id, outfit_snapshot, saved_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        user_id,
                        outfit_id,
                        outfit_snapshots.get(outfit_id, "{}"),
                        (scenario_start + timedelta(minutes=3)).isoformat(),
                    ),
                )
            if user_id % 3 == 0:
                _event(conn, user_id, "show_more", scenario_start + timedelta(minutes=3, seconds=20))
            if user_id % 4 == 0:
                _event(conn, user_id, "reference_opened", scenario_start + timedelta(minutes=3, seconds=40))
            if user_id % 5 == 0:
                _event(conn, user_id, "links_opened", scenario_start + timedelta(minutes=3, seconds=50))
            score = 5 if user_id % 6 not in (0, 1) else (3 if user_id % 6 == 1 else 1)
            _event(conn, user_id, "result_feedback", scenario_start + timedelta(minutes=4), {"score": score})
            if score < 5:
                _event(
                    conn,
                    user_id,
                    "result_feedback_comment",
                    scenario_start + timedelta(minutes=4, seconds=15),
                    {"score": score, "text": "хочется больше визуальных референсов"},
                )

    for ai_index in range(ai_count):
        ai_time = start + timedelta(minutes=2 + ai_index)
        if ai_index % 2 == 0:
            _event(conn, user_id, "item_flow_started", ai_time)
            _event(conn, user_id, "ai_item_outfits_generated", ai_time + timedelta(seconds=20), {"count": 2, "query": "demo"})
        else:
            _event(conn, user_id, "review_started", ai_time)
            _event(conn, user_id, "review_completed", ai_time + timedelta(seconds=25), {"source": "deepseek"})
            _event(conn, user_id, "review_feedback", ai_time + timedelta(seconds=35), {"score": 5 if user_id % 4 else 3})

    if user_id % 4 == 0:
        _event(conn, user_id, "premium_viewed", start + timedelta(minutes=5))
        _event(conn, user_id, "premium_interest", start + timedelta(minutes=5, seconds=10))

    for days, field in [(1, "retention_d1"), (7, "retention_d7"), (30, "retention_d30")]:
        if int(row[field]):
            _event(conn, user_id, "bot_started", start + timedelta(days=days))

    for extra in range(max(0, bot_requests - scenarios - 1)):
        _event(conn, user_id, "bot_started", start + timedelta(minutes=6 + extra))

    _event(
        conn,
        user_id,
        "session_summary",
        end,
        {
            "avg_scenarios": float(row["avg_scenarios"]),
            "time_to_result_min": float(row["time_to_result_min"]),
            "session_length_min": float(row["session_length_min"]),
            "bot_requests_count": int(row["bot_requests_count"]),
            "ai_usage_count": int(row["ai_usage_count"]),
        },
    )


def seed_from_xlsx(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(path)
    init_db()
    close_connection()

    metrics_table: list[list[object]]
    if load_workbook:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook["Dataset"]
        table = [[cell.value for cell in row] for row in sheet.iter_rows()]
        metrics_table = [[cell.value for cell in row] for row in workbook["Metrics"].iter_rows()]
    else:
        table = _read_xlsx_sheet(path, "xl/worksheets/sheet1.xml")
        metrics_table = _read_xlsx_sheet(path, "xl/worksheets/sheet2.xml")

    headers = table[0]
    rows = []
    for raw in table[1:]:
        item = dict(zip(headers, raw))
        rows.append(
            {
                "user_code": str(item["User_ID"]),
                "first_contact_at": _parse_dt(item["Дата обращения к боту"]),
                "scenario_completed": int(item["Конверсия в завершение сценария (1/0)"]),
                "time_to_result_min": float(item["Время до получения результата (мин)"]),
                "avg_scenarios": float(item["Среднее количество сценариев"]),
                "retention_d1": int(item["Retention D1"]),
                "retention_d7": int(item["Retention D7"]),
                "retention_d30": int(item["Retention D30"]),
                "ai_stylist_used": int(item["Использование ИИ-стилиста (1/0)"]),
                "session_length_min": float(item["Длина сессии (мин)"]),
                "bot_requests_count": int(item["Кол-во обращений к боту"]),
                "ai_usage_count": int(item["Кол-во использований ИИ"]),
            }
        )

    db_path = Path(settings.DB_PATH)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("DELETE FROM events WHERE user_id BETWEEN ? AND ?", (DEMO_USER_START, DEMO_USER_END))
    conn.execute("DELETE FROM saved_outfits WHERE user_id BETWEEN ? AND ?", (DEMO_USER_START, DEMO_USER_END))
    conn.execute("DELETE FROM profiles WHERE user_id BETWEEN ? AND ?", (DEMO_USER_START, DEMO_USER_END))
    conn.execute("DELETE FROM ai_quotas WHERE user_id BETWEEN ? AND ?", (DEMO_USER_START, DEMO_USER_END))
    conn.execute("DELETE FROM users WHERE user_id BETWEEN ? AND ?", (DEMO_USER_START, DEMO_USER_END))
    conn.executescript(
        """
        DELETE FROM aarrr_dataset;
        DELETE FROM app_meta;
        """
    )
    conn.execute(
        "INSERT INTO app_meta (key, value) VALUES ('marketing_spend', ?)",
        (str(MARKETING_SPEND_RUB),),
    )
    if len(metrics_table) > 1:
        expected_metrics = {
            str(row[0]): row[1]
            for row in metrics_table[1:]
            if len(row) >= 2 and row[0] not in (None, "")
        }
        conn.execute(
            "INSERT INTO app_meta (key, value) VALUES ('xlsx_metrics', ?)",
            (json.dumps(expected_metrics, ensure_ascii=False),),
        )

    for row in rows:
        conn.execute(
            """
            INSERT INTO aarrr_dataset (
                user_code, first_contact_at, scenario_completed, time_to_result_min,
                avg_scenarios, retention_d1, retention_d7, retention_d30,
                ai_stylist_used, session_length_min, bot_requests_count, ai_usage_count
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["user_code"],
                row["first_contact_at"].isoformat(),
                row["scenario_completed"],
                row["time_to_result_min"],
                row["avg_scenarios"],
                row["retention_d1"],
                row["retention_d7"],
                row["retention_d30"],
                row["ai_stylist_used"],
                row["session_length_min"],
                row["bot_requests_count"],
                row["ai_usage_count"],
            ),
        )

    team_outfits = [outfit for outfit in get_outfits() if outfit.source == "team_curated_txt"]
    outfit_ids = [outfit.id for outfit in team_outfits] or ["demo_outfit"]
    outfit_snapshots = {outfit.id: outfit.model_dump_json() for outfit in team_outfits}
    for index, row in enumerate(rows, start=1):
        demo_user_id = 900000 + index
        _seed_user_events(conn, row, demo_user_id, outfit_ids, outfit_snapshots)
        conn.execute(
            """
            INSERT OR REPLACE INTO ai_quotas (user_id, used_count, limit_count, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (
                demo_user_id,
                min(int(row["ai_usage_count"]), settings.AI_FREE_LIMIT),
                settings.AI_FREE_LIMIT,
                row["first_contact_at"].isoformat(),
            ),
        )

    conn.commit()
    conn.close()
    print(f"Seeded {len(rows)} AARRR users from {path}")


if __name__ == "__main__":
    seed_from_xlsx(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX)
